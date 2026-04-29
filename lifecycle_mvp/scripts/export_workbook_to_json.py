from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\data_lifecycles\Master\Master_Data_Lifecycle_v11_MASTER.xlsx")
OUTPUT = Path(r"D:\Management\Github\dragons8mycat.github.io\lifecycle_mvp\data\mvp-data.json")

ROLE_LABELS = {
    "Analytical": "Critical for decision",
    "Basemapping": "Needed to operate",
    "Descriptive/Contextual": "Helpful context",
    "Unknown": "Needs validation",
}

ROLE_ORDER = ["Analytical", "Basemapping", "Descriptive/Contextual", "Unknown"]


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_key_map(headers: list[str]) -> dict[str, int]:
    return {header: idx for idx, header in enumerate(headers)}


def dataset_status(raw_status: str) -> str:
    status = raw_status.lower()
    if "review" in status:
        return "review"
    if "gap" in status or "no common" in status:
        return "gap"
    return "held"


def dataset_status_label(code: str) -> str:
    return {
        "held": "In hand",
        "review": "Needs validation",
        "gap": "Need to source",
        "unknown": "Unknown role",
    }[code]


def stage_purpose(stage_name: str) -> str:
    purpose_map = {
        "Scoping": "Understand whether the opportunity is worth pursuing and what early risks could change the direction of travel.",
        "Feasibility": "Test whether the project is viable with enough evidence to support investment, screening, and early design confidence.",
        "Preliminary Environmental Screening": "Identify material environmental constraints before detailed assessment work begins.",
        "Environmental Impact Assessment (EIA)": "Build a defensible evidence base for formal assessment, mitigation, and consenting discussions.",
        "Concept Design & Planning  Application": "Support concept definition, site strategy, and a credible planning or consent submission.",
        "Government & Community  Approvals": "Strengthen the case for approvals by surfacing community, policy, and statutory constraints clearly.",
        "Detailed Design & Engineering": "Give engineers and delivery teams the data needed to define the scheme accurately and reduce redesign risk.",
        "Financing &  Acquisition": "Support valuation, acquisition, and commercial confidence with reliable evidence of constraints and opportunity.",
        "Construction": "Reduce delivery surprises by ensuring build teams have the operational and analytical context they need on site.",
        "Sales, Marketing & Handover": "Support handover, customer messaging, and close-out with clear evidence and project context.",
        "Post-Construction  Monitoring": "Track long-term outcomes, compliance, and ongoing performance after delivery.",
        "Strategic Planning (HLP)": "Shape network or site strategy early, using enough evidence to prioritise the right areas and corridors.",
        "High-Level Design (HLD)": "Support route and design choices with the data needed to make credible high-level technical decisions.",
        "Physical Infrastructure Analysis (PIA)": "Assess the physical environment and infrastructure constraints before committing to field work and detailed design.",
        "Field Survey": "Guide survey effort toward the most important gaps, risks, and technical unknowns.",
        "Low Level Design (LLD)": "Complete detailed design with the precision needed for delivery, procurement, and construction readiness.",
        "Civils & Build": "Equip build teams with practical, location-specific data to reduce disruption, risk, and rework.",
        "As-Built": "Capture the final delivered position and supporting context so the asset can be handed over and reused confidently.",
    }
    return purpose_map.get(stage_name, "Use this stage view to understand the data needed to proceed with confidence.")


def readiness_for(counts: dict[str, int], role_counts: Counter[str]) -> tuple[str, str]:
    if counts["gap"] >= 8 or role_counts["Analytical"] > 0 and counts["gap"] >= 3:
        return "blocked", "Material gaps remain at this stage, so the current evidence base is not yet strong enough to proceed confidently."
    if counts["gap"] > 0 or counts["review"] > 0 or counts["unknownRole"] > 0:
        return "at-risk", "Some evidence is in hand, but there are still gaps, unresolved matches, or role uncertainties that should be addressed."
    return "ready", "The current snapshot suggests the core data is in hand for this stage, with no immediate gap or review flags."


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws_master = workbook["1_Data_Master_Expanded"]
    ws_stage = workbook["2_Project_Stage_Data"]
    ws_recon = workbook["0_Reconciliation"]
    ws_additions = workbook["4_Data_Additions"]

    master_rows = list(ws_master.iter_rows(values_only=True))
    stage_rows = list(ws_stage.iter_rows(values_only=True))
    recon_rows = list(ws_recon.iter_rows(values_only=True))
    addition_rows = list(ws_additions.iter_rows(values_only=True))

    master_headers = [clean(value) for value in master_rows[0]]
    stage_headers = [clean(value) for value in stage_rows[0]]
    recon_headers = [clean(value) for value in recon_rows[0]]
    addition_headers = [clean(value) for value in addition_rows[0]]

    master_map = as_key_map(master_headers)
    stage_map = as_key_map(stage_headers)
    recon_map = as_key_map(recon_headers)
    addition_map = as_key_map(addition_headers)

    master_lookup = {}
    usage_by_data_id: dict[str, dict[str, object]] = defaultdict(lambda: {
        "usage_count": 0,
        "project_types": set(),
        "stage_names": set(),
    })
    status_counter: Counter[str] = Counter()

    for row in master_rows[1:]:
        if not row or not row[0]:
            continue
        data_id = clean(row[master_map["data_id"]])
        match_status = clean(row[master_map["match_status"]])
        master_lookup[data_id] = {
            "dataId": data_id,
            "commonName": clean(row[master_map["common_name"]]),
            "dataClass": clean(row[master_map["data_class"]]),
            "productFamily": clean(row[master_map["product_family"]]),
            "supplier": clean(row[master_map["supplier"]]),
            "variantCount": int(row[master_map["variant_count"]] or 0),
            "matchStatus": match_status,
            "statusCode": dataset_status(match_status),
            "confidence": row[master_map["confidence"]],
            "source": clean(row[master_map["source"]]),
        }
        status_counter[match_status] += 1

    review_counter = 0
    review_queue = []
    for row in recon_rows[1:]:
        if not row or not row[0]:
            continue
        match_type = clean(row[recon_map["match_type"]])
        confidence = clean(row[recon_map["confidence"]])
        if match_type in {"review-needed", "no-match"}:
            review_counter += 1
            review_queue.append({
                "queueType": "Catalogue Review",
                "title": clean(row[recon_map["lifecycle_common_name"]]),
                "status": "Needs review",
                "description": f"Match type: {match_type}. Confidence: {confidence or 'N/A'}.",
            })

    project_stage_rows: dict[str, dict[str, list[dict[str, object]]]] = defaultdict(lambda: defaultdict(list))
    project_stage_order: dict[str, dict[str, int]] = defaultdict(dict)
    project_types = set()
    stages = set()

    for row in stage_rows[1:]:
        if not row or not row[0]:
            continue

        project_type = clean(row[stage_map["project_type"]])
        stage_name = clean(row[stage_map["stage_name"]])
        data_id = clean(row[stage_map["data_id"]])
        used_in_stage = clean(row[stage_map["used_in_stage"]])
        stage_role = clean(row[stage_map["stage_data_role"]]) or "Unknown"
        source_record = master_lookup.get(data_id, {
            "dataId": data_id,
            "commonName": clean(row[stage_map["common_name"]]),
            "dataClass": clean(row[stage_map["data_class"]]),
            "productFamily": clean(row[stage_map["product_family"]]),
            "supplier": "",
            "variantCount": 0,
            "matchStatus": "Unknown",
            "statusCode": "held",
            "confidence": "",
            "source": "",
        })

        if data_id:
            usage = usage_by_data_id[data_id]
            usage["usage_count"] += 1
            usage["project_types"].add(project_type)
            usage["stage_names"].add(stage_name)

        project_types.add(project_type)
        stages.add(stage_name)

        if stage_name and stage_name not in project_stage_order[project_type]:
            try:
                stage_sequence = int(row[stage_map["stage_sequence"]] or 999)
            except (ValueError, TypeError):
                stage_sequence = 999
            project_stage_order[project_type][stage_name] = stage_sequence

        if used_in_stage != "Yes":
            continue

        status_code = source_record["statusCode"]
        if stage_role == "Unknown":
            status_for_view = "unknown"
        else:
            status_for_view = status_code

        project_stage_rows[project_type][stage_name].append({
            "dataId": source_record["dataId"],
            "commonName": source_record["commonName"],
            "productFamily": source_record["productFamily"],
            "supplier": source_record["supplier"],
            "source": source_record["source"],
            "role": stage_role,
            "status": status_for_view,
            "statusLabel": dataset_status_label(status_for_view),
            "notes": clean(row[stage_map["notes"]]),
        })

    strategic_gaps_by_project: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in addition_rows[1:]:
        if not row or not row[0]:
            continue
        project_type = clean(row[addition_map["project_type"]])
        strategic_gaps_by_project[project_type].append({
            "commonName": clean(row[addition_map["common_name"]]),
            "source": clean(row[addition_map["source"]]),
            "targetRoles": clean(row[addition_map["target_roles"]]),
        })

    project_type_payload = []
    for project_type in sorted(project_stage_rows):
        sorted_stage_names = sorted(project_stage_rows[project_type], key=lambda name: project_stage_order[project_type].get(name, 999))
        stages_payload = []
        blocked_count = 0
        at_risk_count = 0

        for stage_name in sorted_stage_names:
            records = project_stage_rows[project_type][stage_name]
            role_counts = Counter(record["role"] for record in records)
            counts = {
                "held": sum(1 for record in records if record["status"] == "held"),
                "review": sum(1 for record in records if record["status"] == "review"),
                "gap": sum(1 for record in records if record["status"] == "gap"),
                "unknownRole": sum(1 for record in records if record["status"] == "unknown"),
            }
            readiness, readiness_narrative = readiness_for(counts, role_counts)
            if readiness == "blocked":
                blocked_count += 1
            elif readiness == "at-risk":
                at_risk_count += 1

            grouped = defaultdict(list)
            for record in sorted(records, key=lambda item: (ROLE_ORDER.index(item["role"]) if item["role"] in ROLE_ORDER else 99, item["commonName"])):
                grouped[record["role"]].append(record)

            role_groups = []
            for role in ROLE_ORDER:
                role_groups.append({
                    "key": role,
                    "label": ROLE_LABELS[role],
                    "items": grouped.get(role, []),
                })

            top_held = [record["commonName"] for record in records if record["status"] == "held"][:6]
            top_gaps = [record["commonName"] for record in records if record["status"] in {"gap", "review", "unknown"}][:6]

            stages_payload.append({
                "stageName": stage_name,
                "stagePurpose": stage_purpose(stage_name),
                "readiness": readiness,
                "readinessNarrative": readiness_narrative,
                "counts": counts,
                "roleCounts": {
                    "Analytical": role_counts["Analytical"],
                    "Basemapping": role_counts["Basemapping"],
                    "Descriptive/Contextual": role_counts["Descriptive/Contextual"],
                    "Unknown": role_counts["Unknown"],
                },
                "topHeld": top_held,
                "topGaps": top_gaps,
                "roleGroups": role_groups,
            })

        project_type_payload.append({
            "name": project_type,
            "stageCount": len(stages_payload),
            "atRiskCount": at_risk_count,
            "blockedCount": blocked_count,
            "strategicGapCount": len(strategic_gaps_by_project[project_type]),
            "strategicGaps": strategic_gaps_by_project[project_type],
            "stages": stages_payload,
        })

    datasets = []
    for data_id, master in master_lookup.items():
        usage = usage_by_data_id[data_id]
        datasets.append({
            "dataId": data_id,
            "commonName": master["commonName"],
            "dataClass": master["dataClass"],
            "productFamily": master["productFamily"],
            "supplier": master["supplier"],
            "variantCount": master["variantCount"],
            "matchStatus": master["matchStatus"],
            "confidence": master["confidence"],
            "source": master["source"],
            "usageCount": usage["usage_count"],
            "projectTypes": sorted(item for item in usage["project_types"] if item),
            "stageNames": sorted(item for item in usage["stage_names"] if item),
        })

    datasets.sort(key=lambda item: (-item["usageCount"], item["commonName"]))

    payload = {
        "summary": {
            "datasetCount": len(datasets),
            "stageRowCount": len(stage_rows) - 1,
            "projectTypeCount": len(project_types),
            "reviewCount": review_counter,
        },
        "filters": {
            "projectTypes": sorted(item for item in project_types if item),
            "stages": sorted(item for item in stages if item),
            "matchStatuses": sorted(status for status in status_counter if status),
        },
        "projectTypes": project_type_payload,
        "datasets": datasets,
        "reviewQueue": review_queue[:12],
    }

    OUTPUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
