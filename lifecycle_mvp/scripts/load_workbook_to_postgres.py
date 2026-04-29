from __future__ import annotations

import csv
import os
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\data_lifecycles\Master\Master_Data_Lifecycle_v11_MASTER.xlsx")
SQL_DIR = Path(r"D:\Management\Github\dragons8mycat.github.io\lifecycle_mvp\sql")
PSQL = "psql"
DB_HOST = os.getenv("PGHOST", "db.uzlslubzcfuaynsjiuxj.supabase.co")
DB_PORT = os.getenv("PGPORT", "5432")
DB_NAME = os.getenv("PGDATABASE", "postgres")
DB_USER = os.getenv("PGUSER", "postgres")
SCHEMA = "data_lifecycles_mvp"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def run_psql(command: str, env: dict[str, str]) -> None:
    subprocess.run(
        [PSQL, "-h", DB_HOST, "-p", DB_PORT, "-U", DB_USER, "-d", DB_NAME, "-v", "ON_ERROR_STOP=1", "-c", command],
        check=True,
        env=env,
    )


def run_psql_file(path: Path, env: dict[str, str]) -> None:
    subprocess.run(
        [PSQL, "-h", DB_HOST, "-p", DB_PORT, "-U", DB_USER, "-d", DB_NAME, "-v", "ON_ERROR_STOP=1", "-f", str(path)],
        check=True,
        env=env,
    )


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def main() -> None:
    if "PGPASSWORD" not in os.environ:
        raise SystemExit("PGPASSWORD must be set before running this loader.")

    env = os.environ.copy()

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws_recon = workbook["0_Reconciliation"]
    ws_master = workbook["1_Data_Master_Expanded"]
    ws_stage = workbook["2_Project_Stage_Data"]
    ws_variants = workbook["3_Catalogue_Variants"]
    ws_additions = workbook["4_Data_Additions"]

    master_rows = list(ws_master.iter_rows(values_only=True))
    stage_rows = list(ws_stage.iter_rows(values_only=True))
    recon_rows = list(ws_recon.iter_rows(values_only=True))
    variant_rows = list(ws_variants.iter_rows(values_only=True))
    addition_rows = list(ws_additions.iter_rows(values_only=True))

    master_headers = [clean(item) for item in master_rows[0]]
    stage_headers = [clean(item) for item in stage_rows[0]]
    recon_headers = [clean(item) for item in recon_rows[0]]
    variant_headers = [clean(item) for item in variant_rows[0]]
    addition_headers = [clean(item) for item in addition_rows[0]]

    master_idx = {name: idx for idx, name in enumerate(master_headers)}
    stage_idx = {name: idx for idx, name in enumerate(stage_headers)}
    recon_idx = {name: idx for idx, name in enumerate(recon_headers)}
    variant_idx = {name: idx for idx, name in enumerate(variant_headers)}
    addition_idx = {name: idx for idx, name in enumerate(addition_headers)}

    project_type_rows: list[list[object]] = []
    project_type_map: dict[str, int] = {}
    for row in stage_rows[1:]:
        project_type = clean(row[stage_idx["project_type"]])
        sector_group = clean(row[stage_idx["sector_group"]])
        if not project_type or project_type in project_type_map:
            continue
        project_type_map[project_type] = len(project_type_map) + 1
        project_type_rows.append([project_type_map[project_type], project_type, sector_group])

    stage_rows_out: list[list[object]] = []
    stage_map: dict[tuple[str, str], int] = {}
    for row in stage_rows[1:]:
        project_type = clean(row[stage_idx["project_type"]])
        stage_name = clean(row[stage_idx["stage_name"]])
        key = (project_type, stage_name)
        if not project_type or not stage_name or key in stage_map:
            continue
        stage_id = len(stage_map) + 1
        stage_map[key] = stage_id
        stage_rows_out.append([
            stage_id,
            project_type_map[project_type],
            stage_name,
            row[stage_idx["stage_sequence"]] or None,
        ])

    asset_rows_out: list[list[object]] = []
    asset_id_map: dict[str, int] = {}
    common_name_to_asset_id: dict[str, int] = {}
    for row_number, row in enumerate(master_rows[1:], start=1):
        data_id = clean(row[master_idx["data_id"]])
        if not data_id:
            continue
        asset_id = row_number
        common_name = clean(row[master_idx["common_name"]])
        asset_id_map[data_id] = asset_id
        if common_name and common_name not in common_name_to_asset_id:
            common_name_to_asset_id[common_name] = asset_id
        asset_rows_out.append([
            asset_id,
            data_id,
            common_name,
            clean(row[master_idx["data_class"]]),
            clean(row[master_idx["product_family"]]),
            clean(row[master_idx["iso_category"]]),
            clean(row[master_idx["supplier"]]),
            clean(row[master_idx["open_proprietary"]]),
            clean(row[master_idx["coverage"]]),
            clean(row[master_idx["origin_company"]]),
            clean(row[master_idx["source"]]),
            clean(row[master_idx["match_status"]]),
            clean(row[master_idx["populated_by"]]),
            row[master_idx["confidence"]] or None,
        ])

    recon_rows_out: list[list[object]] = []
    for row_number, row in enumerate(recon_rows[1:], start=1):
        common_name = clean(row[recon_idx["lifecycle_common_name"]])
        recon_rows_out.append([
            row_number,
            common_name_to_asset_id.get(common_name),
            common_name,
            clean(row[recon_idx["catalogue_matches"]]),
            clean(row[recon_idx["match_type"]]),
            row[recon_idx["confidence"]] or None,
            clean(row[recon_idx["populated_by"]]),
            "needs_review",
            "",
        ])

    variant_rows_out: list[list[object]] = []
    for row_number, row in enumerate(variant_rows[1:], start=1):
        common_name = clean(row[variant_idx["common_name"]])
        variant_rows_out.append([
            row_number,
            common_name_to_asset_id.get(common_name),
            clean(row[variant_idx["data_name"]]),
            clean(row[variant_idx["data_name"]]).lower(),
            common_name,
            clean(row[variant_idx["supplier"]]),
            clean(row[variant_idx["product_family"]]),
            clean(row[variant_idx["iso_category"]]),
            clean(row[variant_idx["coverage"]]),
            clean(row[variant_idx["open_proprietary"]]),
            clean(row[variant_idx["origin"]]),
            "",
        ])

    requirement_rows_out: list[list[object]] = []
    for row_number, row in enumerate(stage_rows[1:], start=1):
        project_type = clean(row[stage_idx["project_type"]])
        stage_name = clean(row[stage_idx["stage_name"]])
        data_id = clean(row[stage_idx["data_id"]])
        requirement_rows_out.append([
            row_number,
            clean(row[stage_idx["key"]]),
            project_type_map[project_type],
            stage_map[(project_type, stage_name)],
            asset_id_map.get(data_id),
            data_id,
            clean(row[stage_idx["common_name"]]),
            clean(row[stage_idx["stage_data_role"]]),
            clean(row[stage_idx["role_code"]]),
            clean(row[stage_idx["used_in_stage"]]),
            clean(row[stage_idx["activity"]]),
            clean(row[stage_idx["populated_by"]]),
            row[stage_idx["confidence"]] or None,
            clean(row[stage_idx["notes"]]),
            clean(row[stage_idx["lookup_key"]]),
        ])

    addition_rows_out: list[list[object]] = []
    for row_number, row in enumerate(addition_rows[1:], start=1):
        project_type = clean(row[addition_idx["project_type"]])
        addition_rows_out.append([
            row_number,
            project_type_map.get(project_type),
            clean(row[addition_idx["data_id"]]),
            clean(row[addition_idx["common_name"]]),
            clean(row[addition_idx["data_class"]]),
            clean(row[addition_idx["product_family"]]),
            clean(row[addition_idx["iso_category"]]),
            clean(row[addition_idx["supplier"]]),
            clean(row[addition_idx["open_proprietary"]]),
            clean(row[addition_idx["coverage"]]),
            clean(row[addition_idx["origin_company"]]),
            clean(row[addition_idx["catalogue_variant_names"]]),
            clean(row[addition_idx["target_roles"]]),
            clean(row[addition_idx["source"]]),
        ])

    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        csv_files = {
            "project_type": temp_dir / "project_type.csv",
            "project_stage": temp_dir / "project_stage.csv",
            "asset": temp_dir / "asset.csv",
            "asset_reconciliation": temp_dir / "asset_reconciliation.csv",
            "asset_variant": temp_dir / "asset_variant.csv",
            "stage_asset_requirement": temp_dir / "stage_asset_requirement.csv",
            "manual_addition": temp_dir / "manual_addition.csv",
        }

        write_csv(csv_files["project_type"], ["project_type_id", "name", "sector_group"], project_type_rows)
        write_csv(csv_files["project_stage"], ["stage_id", "project_type_id", "stage_name", "stage_sequence"], stage_rows_out)
        write_csv(csv_files["asset"], ["asset_id", "data_id", "common_name", "data_class", "product_family", "iso_category", "supplier", "open_proprietary", "coverage", "origin_company", "source", "match_status", "populated_by", "confidence"], asset_rows_out)
        write_csv(csv_files["asset_reconciliation"], ["reconciliation_id", "asset_id", "lifecycle_common_name", "catalogue_matches", "match_type", "confidence", "populated_by", "review_status", "notes"], recon_rows_out)
        write_csv(csv_files["asset_variant"], ["variant_id", "asset_id", "raw_name", "normalized_name", "common_name", "supplier", "product_family", "iso_category", "coverage", "open_proprietary", "origin_company", "import_run_id"], variant_rows_out)
        write_csv(csv_files["stage_asset_requirement"], ["requirement_id", "requirement_key", "project_type_id", "stage_id", "asset_id", "data_id", "common_name", "stage_data_role", "role_code", "used_in_stage", "activity", "populated_by", "confidence", "notes", "lookup_key"], requirement_rows_out)
        write_csv(csv_files["manual_addition"], ["addition_id", "project_type_id", "data_id", "common_name", "data_class", "product_family", "iso_category", "supplier", "open_proprietary", "coverage", "origin_company", "catalogue_variant_names", "target_roles", "source"], addition_rows_out)

        run_psql_file(SQL_DIR / "001_schema.sql", env)
        run_psql("truncate table data_lifecycles_mvp.stage_asset_requirement, data_lifecycles_mvp.asset_variant, data_lifecycles_mvp.asset_reconciliation, data_lifecycles_mvp.manual_addition, data_lifecycles_mvp.project_stage, data_lifecycles_mvp.project_type, data_lifecycles_mvp.asset restart identity cascade;", env)

        copy_commands = [
            rf"\copy {SCHEMA}.project_type (project_type_id, name, sector_group) from '{csv_files['project_type']}' with (format csv, header true, encoding 'UTF8')",
            rf"\copy {SCHEMA}.project_stage (stage_id, project_type_id, stage_name, stage_sequence) from '{csv_files['project_stage']}' with (format csv, header true, encoding 'UTF8')",
            rf"\copy {SCHEMA}.asset (asset_id, data_id, common_name, data_class, product_family, iso_category, supplier, open_proprietary, coverage, origin_company, source, match_status, populated_by, confidence) from '{csv_files['asset']}' with (format csv, header true, encoding 'UTF8')",
            rf"\copy {SCHEMA}.asset_reconciliation (reconciliation_id, asset_id, lifecycle_common_name, catalogue_matches, match_type, confidence, populated_by, review_status, notes) from '{csv_files['asset_reconciliation']}' with (format csv, header true, encoding 'UTF8')",
            rf"\copy {SCHEMA}.asset_variant (variant_id, asset_id, raw_name, normalized_name, common_name, supplier, product_family, iso_category, coverage, open_proprietary, origin_company, import_run_id) from '{csv_files['asset_variant']}' with (format csv, header true, encoding 'UTF8', null '')",
            rf"\copy {SCHEMA}.stage_asset_requirement (requirement_id, requirement_key, project_type_id, stage_id, asset_id, data_id, common_name, stage_data_role, role_code, used_in_stage, activity, populated_by, confidence, notes, lookup_key) from '{csv_files['stage_asset_requirement']}' with (format csv, header true, encoding 'UTF8', null '')",
            rf"\copy {SCHEMA}.manual_addition (addition_id, project_type_id, data_id, common_name, data_class, product_family, iso_category, supplier, open_proprietary, coverage, origin_company, catalogue_variant_names, target_roles, source) from '{csv_files['manual_addition']}' with (format csv, header true, encoding 'UTF8', null '')",
        ]

        for command in copy_commands:
            run_psql(command, env)

        run_psql_file(SQL_DIR / "002_views.sql", env)

    print("Loaded workbook data into data_lifecycles_mvp.")


if __name__ == "__main__":
    main()
